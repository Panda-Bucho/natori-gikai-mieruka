# -*- coding: utf-8 -*-
"""全市区町村の人口と議員定数を突合して data/council.json を生成する。

出典(3ソース、work/council_src/ にキャッシュ):
- 人口: 令和7年国勢調査 人口速報集計(2025年10月1日現在)
  e-Stat https://www.e-stat.go.jp/stat-search/file-download?statInfId=000040454825&fileKind=0
- 市・特別区の議員定数: 全国市議会議長会「市議会議員定数に関する調査結果」(令和7年12月31日現在)
  https://www.si-gichokai.jp/research/teisu/__icsFiles/afieldfile/2026/06/04/20260604_teisu_tyousa.xlsx
- 町村の議員定数: 全国町村議会議長会「第71回町村議会実態調査」(令和7年7月1日現在)
  https://www.nactva.gr.jp/html/research/pdf/71_2.xls

使い方: python scripts/build_council_data.py
再実行しても同じ結果になる(キャッシュがあれば再ダウンロードしない)。
年次更新時は work/council_src/ を消して、上記URLを最新版に差し替えて実行する。
"""
import json
import os
import re
import sys
import urllib.request
from datetime import datetime, timezone, timedelta

BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
SRC = os.path.join(BASE, "work", "council_src")
OUT = os.path.join(BASE, "data", "council.json")

FILES = {
    "census2025_prelim.xlsx": "https://www.e-stat.go.jp/stat-search/file-download?statInfId=000040454825&fileKind=0",
    "sigichokai_teisu.xlsx": "https://www.si-gichokai.jp/research/teisu/__icsFiles/afieldfile/2026/06/04/20260604_teisu_tyousa.xlsx",
    "nactva_71_2.xls": "https://www.nactva.gr.jp/html/research/pdf/71_2.xls",
}

TOHOKU = {"青森県", "岩手県", "宮城県", "秋田県", "山形県", "福島県"}


def fetch_sources():
    os.makedirs(SRC, exist_ok=True)
    for name, url in FILES.items():
        path = os.path.join(SRC, name)
        if os.path.exists(path):
            continue
        print(f"download: {name}")
        req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0 (compatible; natori-gikai-mieruka data builder)"})
        with urllib.request.urlopen(req, timeout=60) as res, open(path, "wb") as f:
            f.write(res.read())


def norm(name):
    """突合用の名称正規化(表記ゆれ: ヶ/ケ・ヵ/カ・空白)"""
    return re.sub(r"[\s　]", "", str(name)).replace("ヶ", "ケ").replace("ヵ", "カ")


def load_census():
    """code -> {pref, name, pop, type} (1741市区町村)"""
    import openpyxl
    wb = openpyxl.load_workbook(os.path.join(SRC, "census2025_prelim.xlsx"), read_only=True)
    ws = wb.worksheets[0]
    out = {}
    for row in ws.iter_rows(min_row=14, max_col=4, values_only=True):
        kind, pref_raw, area_raw, pop = row
        kind = str(kind or "").strip()
        if kind not in ("1", "2", "3", "0"):
            continue
        m = re.match(r"(\d{5})_(.+)", str(area_raw or ""))
        if not m:
            continue
        code, name = m.group(1), m.group(2).strip()
        pref = str(pref_raw).split("_", 1)[1].strip()
        if kind == "1":
            if code == "13100":  # 特別区部(集計行)は対象外
                continue
            mtype = "市"  # 政令指定都市(市計)
        elif kind == "2":
            mtype = "市"
        elif kind == "3":
            mtype = "町" if name.endswith("町") else "村"
        else:  # kind == "0": 政令市の行政区(議会なし)は対象外、東京特別区のみ採用
            if not code.startswith("131"):
                continue
            mtype = "特別区"
        out[code] = {"pref": pref, "name": name, "pop": int(pop), "type": mtype}
    wb.close()
    assert len(out) == 1741, f"census municipalities = {len(out)} (expected 1741)"
    return out


def load_city_seats():
    """(pref, 正規化名) -> 条例定数 (全国市議会議長会、815市区)"""
    import openpyxl
    wb = openpyxl.load_workbook(os.path.join(SRC, "sigichokai_teisu.xlsx"), read_only=True)
    ws = wb.worksheets[0]
    out = {}
    for row in ws.iter_rows(min_row=5, max_col=6, values_only=True):
        _, pref, city, _, _, seats = row
        if not pref or not city or seats is None:
            continue
        if not re.fullmatch(r"\d+", str(seats).strip().split(".")[0]):
            continue
        out[(str(pref).strip(), norm(city))] = int(float(seats))
    wb.close()
    assert len(out) >= 810, f"city seats = {len(out)} (expected ~815)"
    return out


def load_town_seats():
    """(pref, 正規化名) -> 定数 (全国町村議会議長会、926町村)"""
    import xlrd
    wb = xlrd.open_workbook(os.path.join(SRC, "nactva_71_2.xls"))
    ws = wb.sheet_by_index(0)
    out = {}
    for r in range(ws.nrows):
        pref = str(ws.cell_value(r, 0)).strip()
        town = str(ws.cell_value(r, 1)).strip()
        seats = ws.cell_value(r, 4)
        if not pref.endswith(("都", "道", "府", "県")) or not town:
            continue
        try:
            seats = int(float(seats))
        except (TypeError, ValueError):
            continue
        if seats <= 0:
            continue
        out[(pref, norm(town))] = seats
    assert len(out) >= 920, f"town seats = {len(out)} (expected ~926)"
    return out


def main():
    fetch_sources()
    census = load_census()
    city_seats = load_city_seats()
    town_seats = load_town_seats()

    municipalities = []
    unmatched = []
    used_city, used_town = set(), set()
    for code in sorted(census):
        c = census[code]
        key = (c["pref"], norm(c["name"]))
        if c["type"] in ("市", "特別区"):
            seats = city_seats.get(key)
            if seats is not None:
                used_city.add(key)
        else:
            seats = town_seats.get(key)
            if seats is not None:
                used_town.add(key)
        if seats is None:
            unmatched.append(f"{c['pref']} {c['name']} ({c['type']})")
            continue
        municipalities.append({
            "code": code,
            "name": c["name"],
            "pref": c["pref"],
            "type": c["type"],
            "pop": c["pop"],
            "seats": seats,
        })

    print(f"matched: {len(municipalities)} / {len(census)}")
    if unmatched:
        print("unmatched (census側):")
        for u in unmatched:
            print("  ", u)
    for label, src, used in (("市区", city_seats, used_city), ("町村", town_seats, used_town)):
        leftover = set(src) - used
        if leftover:
            print(f"unmatched ({label}議長会側):")
            for k in sorted(leftover):
                print("  ", k)

    # 検証
    natori = next(m for m in municipalities if m["code"] == "04207")
    assert natori["name"] == "名取市" and natori["seats"] == 21, natori
    miyagi = [m for m in municipalities if m["pref"] == "宮城県"]
    assert len(miyagi) == 35, f"宮城県 {len(miyagi)}団体 (expected 35)"
    cities = [m for m in municipalities if m["type"] == "市"]
    print(f"内訳: 市 {len(cities)} / 特別区 {sum(1 for m in municipalities if m['type']=='特別区')} / "
          f"町 {sum(1 for m in municipalities if m['type']=='町')} / 村 {sum(1 for m in municipalities if m['type']=='村')}")
    print(f"名取市: 人口 {natori['pop']:,} 定数 {natori['seats']} → {natori['seats']/natori['pop']*1000:.3f}人/千人")

    jst = timezone(timedelta(hours=9))
    data = {
        "generatedAt": datetime.now(jst).isoformat(timespec="seconds"),
        "basisPopulation": "令和7年国勢調査 人口速報集計(2025年10月1日現在の速報値)",
        "basisSeatsCity": "全国市議会議長会「市議会議員定数に関する調査結果」(2025年12月31日現在の条例定数)",
        "basisSeatsTown": "全国町村議会議長会「第71回町村議会実態調査」(2025年7月1日現在の条例定数)",
        "municipalities": municipalities,
    }
    os.makedirs(os.path.dirname(OUT), exist_ok=True)
    with open(OUT, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, separators=(",", ":"))
    size = os.path.getsize(OUT)
    print(f"wrote {OUT} ({len(municipalities)}件, {size/1024:.0f}KB)")


if __name__ == "__main__":
    sys.stdout.reconfigure(encoding="utf-8")
    main()
